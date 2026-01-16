"""
OpenAI service for interacting with Responses API (Conversations)
"""
import openai
import time
import logging
import httpx
import os
from typing import Dict, Optional, List, Any
from config import OPENAI_API_KEY, MAX_TOOL_ITERATIONS, MAX_TOOL_CALLS

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Set OpenAI API key
openai.api_key = OPENAI_API_KEY


class OpenAIService:
    """Service class for OpenAI Responses API operations"""
    
    def __init__(self):
        """Initialize OpenAI service"""
        self.client = openai.OpenAI(api_key=OPENAI_API_KEY)
        logger.info("OpenAI Responses API service initialized")
    
    def create_conversation(self, metadata: Optional[Dict] = None) -> str:
        """
        Create a new conversation for stateful interactions
        
        Args:
            metadata: Optional metadata for the conversation
            
        Returns:
            Conversation ID
        """
        try:
            conversation = self.client.conversations.create(
                metadata=metadata or {}
            )
            logger.info(f"Created conversation: {conversation.id}")
            return conversation.id
        except Exception as e:
            logger.error(f"Error creating conversation: {e}")
            raise

    def _extract_excel_text_for_prompt(self, file_path: str) -> Optional[str]:
        """
        NodeJS parity (scenario #2):
        Extract Excel contents into a plain-text, TSV-like dump that can be appended
        to the user message (similar to Wrike-Showcase-GPT's Excel extractor).
        """
        try:
            ext = os.path.splitext(file_path)[1].lower()
            if ext not in [".xlsx", ".xlsm", ".xltx", ".xltm"]:
                return None

            import openpyxl  # local import to avoid hard dependency issues
            from openpyxl.utils.cell import range_boundaries

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_names = list(wb.sheetnames)

            extracted = "Excel Data Summary:\n\n"
            extracted += f"File: {os.path.basename(file_path)}\n"
            extracted += f"Total Sheets: {len(sheet_names)}\n\n"

            for sheet_name in sheet_names:
                ws = wb[sheet_name]

                # Similar to Node's worksheet['!ref'] used-range.
                dim = None
                try:
                    dim = ws.calculate_dimension()
                except Exception:
                    dim = getattr(ws, "dimensions", None)

                if not dim:
                    min_col, min_row, max_col, max_row = 1, 1, ws.max_column or 1, ws.max_row or 1
                else:
                    try:
                        min_col, min_row, max_col, max_row = range_boundaries(dim)
                    except Exception:
                        min_col, min_row, max_col, max_row = 1, 1, ws.max_column or 1, ws.max_row or 1

                extracted += f"Sheet: {sheet_name}\n"
                extracted += f"Total Rows: {max_row}\n"
                extracted += f"Total Columns: {max_col}\n\n"

                for row in ws.iter_rows(
                    min_row=min_row,
                    max_row=max_row,
                    min_col=min_col,
                    max_col=max_col,
                ):
                    values: List[str] = []
                    any_non_empty = False
                    for cell in row:
                        v = cell.value
                        s = "" if v is None else str(v)
                        s = s.strip()
                        if s:
                            any_non_empty = True
                        values.append(s)
                    if any_non_empty:
                        extracted += "\t".join(values) + "\n"

                extracted += "\n----------------------------------------\n\n"

            return extracted
        except Exception as e:
            logger.warning(f"[EXCEL] Failed to extract Excel text for prompt from {file_path}: {e}")
            return None

    def _estimate_tokens_fast(self, text: str) -> int:
        """
        Very rough token estimate similar to Node's fallback (4 chars per token).
        """
        if not text:
            return 0
        return max(1, len(text) // 4)

    def _get_model_context_window(self, model: str) -> int:
        """
        Lightweight context window mapping (Node has a richer table).
        Keep conservative defaults to avoid truncation.
        """
        m = (model or "").lower()
        # Conservative defaults; adjust if you know exact windows.
        if "gpt-5" in m:
            return 128000
        if "gpt-4" in m:
            return 128000
        return 32000

    def _clip_text_to_token_budget(self, text: str, max_tokens: int) -> str:
        """
        Clip text to approx token budget from the start (same general behavior as Node background limiter).
        """
        if not text:
            return text
        if max_tokens <= 0:
            return ""
        est = self._estimate_tokens_fast(text)
        if est <= max_tokens:
            return text
        # Approx chars to keep
        keep_chars = max(0, max_tokens * 4)
        clipped = text[:keep_chars]
        return clipped + "\n\n[...truncated to fit context window...]\n"
    
    def build_input_from_message(self, 
                                 message_content: str,
                                 file_ids: Optional[List[str]] = None,
                                 tools: Optional[List[Dict]] = None) -> List[Dict]:
        """
        Build input array for Responses API
        
        NodeJS parity (scenario #1):
        - Add {type:'input_file', file_id:'...'} items to the input array so the flow
          mirrors the Node request-building behavior.
        - The final API request will still use code_interpreter container.file_ids.
        
        Args:
            message_content: The message text
            file_ids: Optional list of OpenAI uploaded file IDs
            tools: Unused (kept for backward compatibility with existing callers)
            
        Returns:
            List of input items (message + input_file blocks)
        """
        input_items = []
        
        # Add text message
        if message_content:
            input_items.append({
                "type": "message",
                "role": "user",
                "content": [
                    {
                        "type": "input_text",
                        "text": message_content
                    }
                ]
            })

        # Add input_file blocks (Node style)
        if file_ids:
            for fid in file_ids:
                if fid:
                    input_items.append({"type": "input_file", "file_id": fid})
        
        return input_items
    
    def create_response(self,
                       model: str,
                       instructions: str,
                       input_items: List[Dict],
                       tools: Optional[List[Dict]] = None,
                       conversation_id: Optional[str] = None,
                       max_output_tokens: Optional[int] = None,
                       file_ids: Optional[List[str]] = None,
                       sampling: Optional[Dict] = None) -> Any:
        """
        Create a response using the Responses API
        
        Args:
            model: The model to use
            instructions: System instructions
            input_items: List of input items (messages, files, etc.)
            tools: Optional list of tools
            conversation_id: Optional conversation ID for stateful interactions
            max_output_tokens: Maximum output tokens
            file_ids: Optional list of file IDs for code_interpreter
            sampling: Optional sampling config {temperature, top_p}
            
        Returns:
            Response object
        """
        try:
            # NodeJS parity: derive file IDs from input_file blocks if present
            derived_file_ids: List[str] = []
            if isinstance(input_items, list):
                for it in input_items:
                    if isinstance(it, dict) and it.get("type") == "input_file" and it.get("file_id"):
                        derived_file_ids.append(it["file_id"])

            # Merge explicit file_ids with derived ones (preserve order; de-dupe)
            merged_file_ids: List[str] = []
            for fid in (file_ids or []) + derived_file_ids:
                if fid and fid not in merged_file_ids:
                    merged_file_ids.append(fid)

            request_data = {
                "model": model,
                "instructions": instructions,
                "input": input_items
            }
            
            # DEBUG: Log instructions length and preview
            logger.info(f"[DEBUG] Instructions length: {len(instructions) if instructions else 0} chars")
            logger.info(f"[DEBUG] Instructions preview: {instructions[:200] if instructions else 'None'}...")
            
            # Add sampling (temperature, top_p) if provided - matching MongoDB version
            if sampling:
                if 'temperature' in sampling and sampling['temperature'] is not None:
                    request_data['temperature'] = sampling['temperature']
                if 'top_p' in sampling and sampling['top_p'] is not None:
                    request_data['top_p'] = sampling['top_p']
            
            # Add conversation if provided (stateful mode)
            if conversation_id:
                request_data["conversation"] = conversation_id
                logger.info(f"Using conversation: {conversation_id}")
            else:
                logger.info("Running stateless (no conversation)")
            
            # Add tools if provided
            if tools:
                logger.info(f"[DEBUG] Raw tools input: {tools}")
                # Configure tools with proper format for Responses API
                configured_tools = []
                for tool in tools:
                    if tool.get("type") == "code_interpreter":
                        # Add container configuration for code_interpreter
                        # Files are passed via container.file_ids (NOT in input)
                        configured_tool = {
                            **tool,
                            "container": {
                                "type": "auto",
                                "file_ids": merged_file_ids if merged_file_ids else []
                            }
                        }
                        configured_tools.append(configured_tool)
                    elif tool.get("type") == "file_search":
                        # file_search requires vector_store_ids
                        # If not provided, filter it out to avoid errors
                        logger.warning("file_search tool requires vector_store_ids but none provided. Skipping file_search.")
                        continue
                    elif tool.get("type") in ["web_search", "computer_use", "image_generation"]:
                        # Other built-in tools
                        configured_tools.append(tool)
                    elif tool.get("type") == "function":
                        # Custom function tools
                        configured_tools.append(tool)
                
                if configured_tools:
                    request_data["tools"] = configured_tools

                # NodeJS parity: do not send input_file blocks to the API; only message blocks.
                # Files remain accessible via code_interpreter container.file_ids.
                if isinstance(request_data.get("input"), list):
                    msg_items = [
                        it for it in request_data["input"]
                        if isinstance(it, dict) and it.get("type") == "message"
                    ]
                    if msg_items:
                        request_data["input"] = msg_items
            
                # Add max_output_tokens if specified
                if max_output_tokens:
                    request_data["max_output_tokens"] = max_output_tokens
                
                # DEBUG: Log final request structure (without full instructions)
                debug_request = {k: v for k, v in request_data.items() if k != 'instructions'}
                debug_request['instructions_length'] = len(request_data.get('instructions', ''))
                logger.info(f"[DEBUG] Final request structure: {debug_request}")
                
                # VERIFY: Instructions contain the full Line→Item mapping
                instructions_text = request_data.get('instructions', '')
                logger.info(f"✅ Sending {len(instructions_text)} chars of instructions to OpenAI")
                logger.info(f"✅ Last 400 chars (should show Line→Item mapping): ...{instructions_text[-400:]}")
                
                logger.info(f"Creating response with model: {model}")
                response = self.client.responses.create(**request_data)
            
            # Wait for response to be ready
            response = self._wait_for_response_ready(response)
            
            return response
            
        except Exception as e:
            logger.error(f"Error creating response: {e}")
            raise
    
    def _wait_for_response_ready(self, response: Any, max_wait: int = 300) -> Any:
        """
        Wait for response to be ready (if it's being processed)
        
        Args:
            response: Response object
            max_wait: Maximum seconds to wait
            
        Returns:
            Ready response object
        """
        start_time = time.time()
        while hasattr(response, 'status') and response.status in ['queued', 'in_progress']:
            if time.time() - start_time > max_wait:
                raise TimeoutError(f"Response did not complete within {max_wait} seconds")
            
            logger.info(f"Response status: {response.status}, waiting...")
            time.sleep(2)
            
            # Re-fetch response status if needed
            try:
                response = self.client.responses.retrieve(response.id)
            except:
                # If retrieve not available, continue with existing response
                pass
        
        logger.info(f"Response ready with status: {getattr(response, 'status', 'completed')}")
        return response
    
    def extract_files_from_response(self, response: Any) -> List[Dict]:
        """
        Extract file references from response annotations
        
        Args:
            response: Response object from OpenAI
            
        Returns:
            List of file dictionaries with file_id and filename
        """
        files = []
        
        if hasattr(response, 'output') and isinstance(response.output, list):
            for item in response.output:
                if hasattr(item, 'type') and item.type == 'message':
                    if hasattr(item, 'content'):
                        for content in item.content:
                            if hasattr(content, 'annotations') and content.annotations:
                                for annotation in content.annotations:
                                    if hasattr(annotation, 'type') and annotation.type == 'container_file_citation':
                                        files.append({
                                            'file_id': annotation.file_id,
                                            'filename': annotation.filename,
                                            'container_id': annotation.container_id
                                        })
        
        return files
    
    def delete_file(self, file_id: str) -> bool:
        """
        Delete a file from OpenAI Files storage
        
        Args:
            file_id: The file ID to delete
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # Only delete regular files (not container files)
            if file_id.startswith("cfile_"):
                logger.info(f"Skipping deletion of container file {file_id} (auto-cleaned by OpenAI)")
                return True
            
            logger.info(f"Deleting file {file_id} from OpenAI storage")
            self.client.files.delete(file_id)
            logger.info(f"Successfully deleted file {file_id}")
            return True
            
        except Exception as e:
            logger.error(f"Error deleting file {file_id}: {e}")
            return False
    
    def download_file(self, file_id: str, output_path: str, container_id: Optional[str] = None) -> bool:
        """
        Download a file from OpenAI (regular file or container file)
        
        Args:
            file_id: The file ID to download
            output_path: Local path to save the file
            container_id: Container ID if this is a container file
            
        Returns:
            True if successful, False otherwise
        """
        try:
            logger.info(f"Downloading file {file_id} to {output_path}")
            
            # Check if this is a container file (starts with "cfile_")
            if file_id.startswith("cfile_"):
                if not container_id:
                    logger.error(f"Container file {file_id} requires container_id")
                    return False
                
                # Download container file using containers API via httpx
                try:
                    url = f"https://api.openai.com/v1/containers/{container_id}/files/{file_id}/content"
                    headers = {
                        "Authorization": f"Bearer {OPENAI_API_KEY}",
                        "OpenAI-Beta": "containers=v1"
                    }
                    
                    with httpx.Client(timeout=60.0) as client:
                        response = client.get(url, headers=headers)
                        
                        if response.status_code == 200:
                            # Write to local file
                            with open(output_path, 'wb') as f:
                                f.write(response.content)
                            
                            logger.info(f"Successfully downloaded container file to {output_path}")
                            return True
                        else:
                            logger.error(f"Failed to download container file: HTTP {response.status_code}")
                            logger.error(f"Response: {response.text}")
                            logger.info("Note: Container files are temporary and exist only in the code_interpreter sandbox")
                            logger.info(f"File path in sandbox: /mnt/data/{file_id.replace('cfile_', '')}-{output_path}")
                            return False
                    
                except Exception as container_error:
                    logger.error(f"Error downloading container file: {container_error}")
                    logger.info("Container files may only be accessible through the sandbox environment")
                    return False
            else:
                # Regular file download
                file_content = self.client.files.content(file_id)
                
                # Write to local file
                with open(output_path, 'wb') as f:
                    f.write(file_content.read())
                
                logger.info(f"Successfully downloaded file to {output_path}")
                return True
            
        except Exception as e:
            logger.error(f"Error downloading file {file_id}: {e}")
            return False
    
    def extract_text_from_response(self, response: Any) -> str:
        """
        Extract text content from response output
        
        Args:
            response: Response object from OpenAI
            
        Returns:
            Extracted text
        """
        text_parts = []
        
        # Debug: log response structure
        logger.info(f"Response type: {type(response)}")
        logger.info(f"Response has output: {hasattr(response, 'output')}")
        if hasattr(response, 'output'):
            logger.info(f"Output type: {type(response.output)}")
            logger.info(f"Output length: {len(response.output) if isinstance(response.output, list) else 'N/A'}")
            if isinstance(response.output, list) and len(response.output) > 0:
                logger.info(f"First output item type: {type(response.output[0])}")
                logger.info(f"First output item: {response.output[0]}")
        
        if hasattr(response, 'output') and isinstance(response.output, list):
            for i, item in enumerate(response.output):
                logger.info(f"Processing output item {i}: type={type(item)}")
                
                # Handle array items (OpenAI may return nested arrays)
                if isinstance(item, list):
                    for sub_item in item:
                        if hasattr(sub_item, 'type'):
                            logger.info(f"  Sub-item type: {sub_item.type}")
                            if sub_item.type == 'message':
                                if hasattr(sub_item, 'content'):
                                    for content in sub_item.content:
                                        if hasattr(content, 'type') and content.type in ['text', 'output_text']:
                                            text_parts.append(content.text)
                # Handle direct items
                elif hasattr(item, 'type'):
                    logger.info(f"  Item type: {item.type}")
                    if item.type == 'message':
                        if hasattr(item, 'content'):
                            logger.info(f"  Message content: {item.content}")
                            for content in item.content:
                                if hasattr(content, 'type'):
                                    logger.info(f"    Content type: {content.type}")
                                    if content.type in ['text', 'output_text']:
                                        logger.info(f"    Found text: {content.text[:100]}...")
                                        text_parts.append(content.text)
        
        logger.info(f"Extracted {len(text_parts)} text parts")
        return "\n".join(text_parts)
    
    def upload_file(self, file_path: str, purpose: str = "assistants") -> str:
        """
        Upload a file to OpenAI
        
        Args:
            file_path: Path to the file
            purpose: Purpose of the file (default: "assistants")
            
        Returns:
            File ID
        """
        try:
            # Helpful for debugging: confirm exactly which local files are being attached.
            try:
                size_bytes = os.path.getsize(file_path)
                logger.info(f"Uploading local file: {file_path} ({size_bytes} bytes)")
            except Exception:
                logger.info(f"Uploading local file: {file_path}")

            with open(file_path, "rb") as file:
                response = self.client.files.create(
                    file=file,
                    purpose=purpose
                )
            logger.info(f"Uploaded file: {response.id}")
            return response.id
        except Exception as e:
            logger.error(f"Error uploading file: {e}")
            raise
    
    def add_message_to_conversation(self,
                                    conversation_id: str,
                                    model: str,
                                    instructions: str,
                                    message_content: str,
                                    tools: Optional[List[Dict]] = None,
                                    file_ids: Optional[List[str]] = None) -> Dict:
        """
        Add a message to an existing conversation and get response
        
        Args:
            conversation_id: The conversation ID
            model: The model to use
            instructions: System instructions
            message_content: User message
            tools: Optional tools
            file_ids: Optional file IDs
            
        Returns:
            Dictionary with response text and metadata
        """
        try:
            logger.info(f"Adding message to conversation {conversation_id}")
            
            # Build input
            input_items = self.build_input_from_message(message_content, file_ids, tools)
            
            # Create response in conversation context
            response = self.create_response(
                model=model,
                instructions=instructions,
                input_items=input_items,
                tools=tools,
                conversation_id=conversation_id,
                file_ids=file_ids
            )
            
            # Extract text
            response_text = self.extract_text_from_response(response)
            
            return {
                "response_id": response.id,
                "conversation_id": conversation_id,
                "text": response_text,
                "status": getattr(response, 'status', 'completed'),
                "raw_response": response
            }
            
        except Exception as e:
            logger.error(f"Error adding message to conversation: {e}")
            raise
    
    def get_assistant_response(self,
                                  model: str,
                                  instructions: str,
                                  user_message: str,
                                  tools: Optional[List[Dict]] = None,
                                  file_paths: Optional[List[str]] = None,
                                  use_conversation: bool = False,
                                  conversation_id: Optional[str] = None,
                                  sampling: Optional[Dict] = None,
                                  metadata: Optional[Dict] = None,
                                  output_dir: Optional[str] = None) -> Dict:
        """
        Complete workflow: Send message and get response using Responses API
        
        Args:
            model: The model to use
            instructions: System instructions
            user_message: User's message
            tools: Optional list of tools
            file_paths: Optional file paths to upload
            use_conversation: Whether to use conversation (stateful)
            conversation_id: Existing conversation ID (if continuing)
            sampling: Optional sampling config {temperature, top_p}
            metadata: Optional metadata for new conversations
            
        Returns:
            Dictionary with response and metadata
        """
        try:
            # Upload files if provided
            file_ids = []
            if file_paths:
                for file_path in file_paths:
                    file_id = self.upload_file(file_path)
                    file_ids.append(file_id)
            
            # Create or use conversation
            conv_id = conversation_id
            if use_conversation and not conv_id:
                conv_id = self.create_conversation(metadata=metadata)
            
            # Build input
            # NodeJS parity: append extracted Excel content to the user message
            enhanced_user_message = user_message or ""
            if file_paths:
                for fp in file_paths:
                    excel_text = self._extract_excel_text_for_prompt(fp)
                    if excel_text:
                        enhanced_user_message += f"\n\nExcel Content from {os.path.basename(fp)}:\n{excel_text}\n"

            # NodeJS parity: background limiter (clip injected background to fit context window)
            try:
                context_window = self._get_model_context_window(model)
                completion_budget = int(os.getenv("OPENAI_COMPLETION_BUDGET_TOKENS", "1024"))
                safety_tokens = int(os.getenv("OPENAI_INPUT_SAFETY_TOKENS", "512"))
                trim_margin = int(os.getenv("OPENAI_TRIM_MARGIN_TOKENS", "512"))
                allowed_input = max(1024, context_window - completion_budget - safety_tokens)

                # We can't precisely account for tool schemas; keep extra margin like Node.
                max_bg_tokens = max(0, allowed_input - trim_margin)
                before_tokens = self._estimate_tokens_fast(enhanced_user_message)
                if before_tokens > max_bg_tokens:
                    logger.info(
                        f"[LIMITER] Trimming user message from ~{before_tokens} tokens to ~{max_bg_tokens} tokens to fit context window"
                    )
                    enhanced_user_message = self._clip_text_to_token_budget(enhanced_user_message, max_bg_tokens)
            except Exception as _e:
                pass

            input_items = self.build_input_from_message(
                enhanced_user_message, 
                file_ids if file_ids else None,
                tools
            )
            
            # Create response with sampling
            response = self.create_response(
                model=model,
                instructions=instructions,
                input_items=input_items,
                tools=tools,
                conversation_id=conv_id,
                file_ids=file_ids,
                sampling=sampling  # Pass sampling to match MongoDB version
            )
            
            # Extract text
            response_text = self.extract_text_from_response(response)
            
            # Extract and download files
            files = self.extract_files_from_response(response)
            downloaded_files = []
            
            if files:
                logger.info(f"Found {len(files)} file(s) in response")
                for file_info in files:
                    # Save downloads into output_dir if provided; otherwise current working directory.
                    # Always sanitize the filename to avoid path traversal.
                    safe_filename = os.path.basename(file_info.get('filename') or "")
                    if not safe_filename:
                        safe_filename = f"output_{file_info.get('file_id', 'file')}.bin"

                    if output_dir:
                        os.makedirs(output_dir, exist_ok=True)
                        output_path = os.path.join(output_dir, safe_filename)
                    else:
                        output_path = safe_filename

                    container_id = file_info.get('container_id')
                    if self.download_file(file_info['file_id'], output_path, container_id):
                        downloaded_files.append({
                            'file_id': file_info['file_id'],
                            'filename': file_info['filename'],
                            'local_path': output_path
                        })
            
            # Clean up uploaded input files
            if file_ids:
                logger.info(f"Cleaning up {len(file_ids)} uploaded file(s)")
                for file_id in file_ids:
                    self.delete_file(file_id)
            
            return {
                "response_id": response.id,
                "conversation_id": conv_id,
                "text": response_text,
                "status": getattr(response, 'status', 'completed'),
                "model": model,
                "files": downloaded_files,
                "raw_response": response
            }
            
        except Exception as e:
            logger.error(f"Error in get_assistant_response: {e}")
            raise

